# install package pip install openpyxl

import openpyxl

book = openpyxl.load_workbook("C:\\Users\\raooosub\\PycharmProjects\\PythonProject\\SeleniumExamples\\ExcelDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=3, column=2) # row and column starts from index 1
print(cell.value)

# to write back to excel sheet
sheet.cell(row=3, column=2).value = "James"

cell = sheet.cell(row=3, column=2) # row and column starts from index 1
print(cell.value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['B3'].value)

#To print all data from excel
print("All data from excel")
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

# To print only TC2 values
print("TC2 values only")
for i in range(1, sheet.max_row+1):
    if sheet.cell(i, column=1).value== "TC2":
        for j in range(2, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)


# To print the excel values in dictiionary
Dict= {}

for i in range(1, sheet.max_row+1):
    if sheet.cell(i, column=1).value== "TC2":
        for j in range(2, sheet.max_column+1):
           Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)